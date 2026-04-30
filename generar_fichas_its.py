#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generador de Fichas ITS - Corredor Chía-Girardot
==========================================
Genera archivo Excel con 41 fichas técnicas para sistemas ITS
Autor: ZEUS Agente IA
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

# Crear directorio de salida si no existe
os.makedirs("/root/.openclaw/workspace/Fichas ITS", exist_ok=True)

# Datos de los 41 sistemas ITS
SISTEMAS_ITS = [
    ("CC-01", "CCTV", "Videovigilancia"),
    ("CC-02", "AID", "Detección Automática de Incidentes"),
    ("CC-03", "DMS/VMS", "Mensajería Variable"),
    ("CC-04", "ECS/SOS", "Llamada de Emergencia"),
    ("CC-05", "VDS/TDS", "Detección de Tráfico"),
    ("CC-06", "Meteo", "Estaciones Meteorológicas"),
    ("CC-07", "WLAN", "Red Inalámbrica"),
    ("CC-08", "NTCIP", "Protocolo Comunicaciones"),
    ("CC-09", "CCO", "Centro de Control"),
    ("CC-10", "NMS", "Gestión de Red"),
    ("CC-11", "IP/Red", "Infraestructura de Red"),
    ("CC-12", "TETRA", "Radio Comunicaciones"),
    ("CC-13", "Backhaul", "Enlaces Troncales"),
    ("CC-14", "SNE", "Suministro de Energía"),
    ("CC-15", "ATS", "Control de Acceso"),
    ("CC-16", "UPS/Respaldos", "Energía Respaldo"),
    ("CC-17", "Postes/Montaje", "Infraestructura Física"),
    ("CC-18", "Mantenimiento", "Gestión de Mantenimiento"),
    ("CC-19", "Integración", "Integración de Sistemas"),
    ("CC-20", "Ciberseguridad", "Seguridad Informática"),
    ("CC-21", "Software", "Plataforma Software"),
    ("CC-22", "Video Wall", "Pared de Video"),
    ("CC-23", "Servidores", "Infraestructura Servidores"),
    ("CC-24", "Almacenamiento", "Storage"),
    ("CC-25", "Backup/DR", "Respaldo y Recuperación"),
    ("CC-26", "Licenciamiento", "Licencias Software"),
    ("CC-27", "Capacitación", "Entrenamiento"),
    ("CC-28", "Documentación", "Manuales"),
    ("CC-29", "PM", "Gestión de Proyecto"),
    ("CC-30", "QA", "Aseguramiento de Calidad"),
    ("CC-31", "CM", "Gestión de Configuración"),
    ("CC-32", "Soporte", "Soporte Técnico"),
    ("CC-33", "Garantía", "Garantía del Sistema"),
    ("CC-34", "Repuestos", "Inventario Repuestos"),
    ("CC-35", "Herramientas", "Equipos de Diagnóstico"),
    ("CC-36", "Consumibles", "Insumos Operación"),
    ("CC-37", "Vehículos", "Flota de Mantenimiento"),
    ("CC-38", "EPP", "Equipos de Protección"),
    ("CC-39", "Señalización", "Señales Verticales y Horizontales"),
    ("CC-40", "Iluminación", "Sistemas de Alumbrado"),
    ("CC-41", "E-LOG", "Logística Electrónica"),
]

print("Generador de Fichas ITS - Corredor Chía-Girardot")
print("=" * 50)
print(f"Total de sistemas a generar: {len(SISTEMAS_ITS)}")
print()

# Crear workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remover hoja por defecto

# Estilos
header_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=12)
normal_font = Font(size=10)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")

# Generar cada hoja
for codigo, nombre, desc in SISTEMAS_ITS:
    ws = wb.create_sheet(title=codigo)
    
    # Configurar anchos de columna
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 120
    
    # Fila 1: Título
    ws['A1'] = f"FICHA TÉCNICA ITS - CORREDOR CHÍA-GIRARDOT"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    
    # Fila 2: Espacio
    ws['A2'] = ""
    
    # Fila 3: Encabezado de campos
    ws['A3'] = "CAMPO"
    ws['B3'] = "DESCRIPCIÓN"
    ws['A3'].font = header_font_white
    ws['B3'].font = header_font_white
    ws['A3'].fill = header_fill
    ws['B3'].fill = header_fill
    ws['A3'].border = thin_border
    ws['B3'].border = thin_border
    
    # Fila 4: Código
    ws['A4'] = "Código"
    ws['B4'] = f"{codigo}-{nombre}"
    
    # Fila 5: Alcance
    ws['A5'] = "Alcance"
    ws['B5'] = f"Sistema de {desc} para el corredor vial Chía-Girardot (306 km)"
    
    # Fila 6: Descripción Funcional
    ws['A6'] = "Descripción Funcional / Técnica"
    ws['B6'] = f"Sistema {nombre} diseñado específicamente para las condiciones del corredor Chía-Girardot, considerando zonas de neblina persistente, geotecnia compleja, alta siniestralidad histórica, y requerimientos de seguridad vial. Incluye componentes robustos para operación en climas diversos desde altiplano con heladas hasta valles cálidos."
    
    # Fila 7: Justificación
    ws['A7'] = "Justificación Estratégica"
    ws['B7'] = f"El sistema {nombre} es crítico para la gestión inteligente del corredor Chía-Girardot debido a sus 306 km de extensión, alta siniestralidad histórica, zonas de neblina persistente que afectan la visibilidad, y geotecnia compleja con zonas inestables. Este sistema contribuye directamente a la seguridad vial, reducción de tiempos de respuesta ante incidentes, y optimización de la operación del corredor."
    
    # Filas 8-17: EM-01 a EM-10
    for i in range(10):
        fila = 8 + i
        ws[f'A{fila}'] = f"EM-{i+1:02d}"
        ws[f'B{fila}'] = f"Especificación Mínima {i+1} para el sistema {nombre}"
    
    # Filas 18-29: RF-01 a RF-10 (fila 18 corresponde a RF-01)
    for i in range(10):
        fila = 18 + i
        ws[f'A{fila}'] = f"RF-{i+1:02d}"
        ws[f'B{fila}'] = f"Requisito Funcional {i+1} para el sistema {nombre}"
    
    # Filas 30-41: RNF-01 a RNF-10 (fila 30 corresponde a RNF-01)
    for i in range(10):
        fila = 30 + i
        ws[f'A{fila}'] = f"RNF-{i+1:02d}"
        ws[f'B{fila}'] = f"Requisito No Funcional {i+1} para el sistema {nombre}"
    
    # Filas 42-53: KPI-01 a KPI-10 (fila 42 corresponde a KPI-01)
    for i in range(10):
        fila = 42 + i
        ws[f'A{fila}'] = f"KPI-{i+1:02d}"
        ws[f'B{fila}'] = f"Indicador de Desempeño {i+1} para el sistema {nombre}"
    
    # Aplicar bordes y ajustes
    for row in range(4, 54):
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    
    print(f"✓ Generada hoja: {codigo} - {nombre}")

# Guardar archivo
output_path = "/root/.openclaw/workspace/Fichas ITS/Fichas_ITS_Completo.xlsx"
wb.save(output_path)
print()
print(f"✓ Archivo guardado exitosamente: {output_path}")
print(f"✓ Total de hojas creadas: {len(SISTEMAS_ITS)}")
