#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

os.makedirs("/root/.openclaw/workspace/Fichas ITS", exist_ok=True)

SISTEMAS = [
    ("CC-01", "CCTV", "Videovigilancia"),
    ("CC-02", "AID", "Detección Automática de Incidentes"),
    ("CC-03", "DMS/VMS", "Mensajería Variable"),
    ("CC-04", "ECS/SOS", "Llamada de Emergencia"),
    ("CC-05", "VDS/TDS", "Detección de Tráfico"),
    ("CC-06", "Meteo", "Estaciones Meteorológicas"),
    ("CC-07", "WLAN", "Red Inalámbrica"),
    ("CC-08", "NTCIP", "Protocolo Comunicaciones"),
    ("CC-09", "CCO", "Centro de Control"),
    ("CC-10", "NMS", "Gestión de Red"),
    ("CC-11", "IP/Red", "Infraestructura de Red"),
    ("CC-12", "TETRA", "Radio Comunicaciones"),
    ("CC-13", "Backhaul", "Enlaces Troncales"),
    ("CC-14", "SNE", "Suministro de Energía"),
    ("CC-15", "ATS", "Control de Acceso"),
    ("CC-16", "UPS/Respaldos", "Energía Respaldo"),
    ("CC-17", "Postes/Montaje", "Infraestructura Física"),
    ("CC-18", "Mantenimiento", "Gestión de Mantenimiento"),
    ("CC-19", "Integración", "Integración de Sistemas"),
    ("CC-20", "Ciberseguridad", "Seguridad Informática"),
    ("CC-21", "Software", "Plataforma Software"),
    ("CC-22", "Video Wall", "Pared de Video"),
    ("CC-23", "Servidores", "Infraestructura Servidores"),
    ("CC-24", "Almacenamiento", "Storage"),
    ("CC-25", "Backup/DR", "Respaldo y Recuperación"),
    ("CC-26", "Licenciamiento", "Licencias Software"),
    ("CC-27", "Capacitación", "Entrenamiento"),
    ("CC-28", "Documentación", "Manuales"),
    ("CC-29", "PM", "Gestión de Proyecto"),
    ("CC-30", "QA", "Aseguramiento de Calidad"),
    ("CC-31", "CM", "Gestión de Configuración"),
    ("CC-32", "Soporte", "Soporte Técnico"),
    ("CC-33", "Garantía", "Garantía del Sistema"),
    ("CC-34", "Repuestos", "Inventario Repuestos"),
    ("CC-35", "Herramientas", "Equipos de Diagnóstico"),
    ("CC-36", "Consumibles", "Insumos Operación"),
    ("CC-37", "Vehículos", "Flota de Mantenimiento"),
    ("CC-38", "EPP", "Equipos de Protección"),
    ("CC-39", "Señalización", "Señales Verticales y Horizontales"),
    ("CC-40", "Iluminación", "Sistemas de Alumbrado"),
    ("CC-41", "E-LOG", "Logística Electrónica"),
]

# Contenido técnico detallado
CONTENIDO = {
    "CC-01": {
        "alcance": "Diseño, suministro, instalación, configuración y puesta en marcha del sistema de videovigilancia CCTV para las 306 km del corredor vial Chía-Girardot, incluyendo cámaras de alta definición, infraestructura de transmisión, centrales de video y almacenamiento, con cobertura completa de zonas críticas, puentes y sectores de alta siniestralidad.",
        "descripcion": "Sistema integral de videovigilancia que utiliza cámaras IP de alta resolución (4K) con capacidad de visión nocturna, analítica de video embebida y transmisión redundante. Las cámaras se ubican estratégicamente cada 2 km en tramos rectos y cada 500 m en curvas, puentes y zonas de neblina persistente. Incluye cámaras PTZ para seguimiento de incidentes, cámaras fijas con lentes gran angular, y cámaras térmicas para zonas con baja visibilidad. La transmisión se realiza mediante fibra óptica con respaldo inalámbrico 5G/4G.",
        "justificacion": "El corredor Chía-Girardot presenta alta siniestralidad histórica, zonas de neblina persistente que reducen la visibilidad hasta 30 metros, y geotecnia compleja con zonas inestables. La videovigilancia permite detección temprana de incidentes, respuesta rápida de equipos de emergencia, análisis forense post-incidente y monitoreo continuo del estado de la infraestructura vial.",
        "em": ["Resolución mínima 4K (3840x2160) para cámaras principales y 2K para secundarias, con capacidad de zoom digital sin pérdida de calidad","Visión nocturna mediante sensores de baja iluminación y LED IR con alcance mínimo de 150 metros en oscuridad total","Grado de protección IP66/IP67 para soportar lluvias intensas, polvo y temperaturas entre -5°C y 45°C","Compresión de video H.265+/H.264 con soporte para análisis de video embebido (vehículos detenidos, contravía, peatones)","Redundancia de transmisión: fibra óptica principal con respaldo inalámbrico 5G/4G automático en caso de falla","Tiempo de almacenamiento mínimo de 30 días continuo a máxima resolución con capacidad de expansión a 90 días","Disponibilidad del sistema superior al 99.5% con acuerdos de nivel de servicio garantizados","Integración con protocolos NTCIP para interoperabilidad con otros sistemas ITS del corredor","Cámaras PTZ con velocidad de rotación mínima 180°/segundo y zoom óptico 30x para seguimiento de incidentes","Certificación eléctrica y de seguridad según normas IEC, NTC e ISO aplicables a infraestructura vial"],
        "rf": ["El sistema debe permitir visualización en tiempo real de todas las cámaras desde el CCO con latencia inferior a 500ms","Debe generar alertas automáticas ante eventos predefinidos: vehículo detenido, peatón en vía, circulación en contravía, congestión","Capacidad de exportación de video con metadata completa (fecha, hora, ubicación GPS, ID de cámara) para análisis forense","Interfaz de usuario intuitiva con mapa georeferenciado del corredor mostrando ubicación de cámaras y su estado operativo","Sistema de gestión de usuarios con roles diferenciados: operador, supervisor, administrador y auditor","Funcionalidad de búsqueda retrospectiva por eventos, fechas, zonas geográficas y características del objeto detectado","Integración bidireccional con el sistema AID para confirmación visual de incidentes detectados automáticamente","Capacidad de visualización remota segura mediante aplicación móvil para personal autorizado con autenticación multifactor","Generación automática de reportes de disponibilidad, eventos detectados y uso del sistema con programación configurable","Soporte para visualización en múltiples pantallas simultáneas con layouts configurables según el rol del operador"],
        "rnf": ["Tiempo de respuesta de la interfaz de usuario inferior a 2 segundos para cualquier operación de navegación entre cámaras","Capacidad del sistema para soportar 200 usuarios concurrentes sin degradación del servicio","Cifrado de transmisión TLS 1.3 para todo el tráfico de video y control entre cámaras y centrales","Arquitectura escalable que permita incorporar nuevas cámaras sin interrupción del servicio operativo","Tiempo medio entre fallos (MTBF) superior a 50,000 horas para componentes críticos","Tiempo medio de reparación (MTTR) inferior a 4 horas con repuestos disponibles en bodega regional","Cumplimiento con estándares de ciberseguridad ISO 27001 y guías del Ministerio de Transporte para sistemas ITS","Interoperabilidad garantizada con sistemas de terceros mediante APIs REST documentadas y protocolos estándar","Documentación técnica completa incluyendo diagramas de red, configuraciones y procedimientos operativos","Soporte técnico 24/7/365 con tiempo de respuesta garantizado según severidad del incidente"],
        "kpi": ["Disponibilidad del sistema CCTV: porcentaje de cámaras operativas en tiempo real, meta superior al 99%","Tiempo promedio de detección de incidentes desde la ocurrencia hasta la alerta en el CCO, meta inferior a 30 segundos","Tiempo de respuesta del equipo de emergencia desde la alerta del CCTV hasta llegada al sitio del incidente, meta inferior a 15 minutos","Número de incidentes detectados automáticamente versus detectados manualmente, meta mayor al 85% de detección automática","Calidad de video medida en términos de resolución efectiva y ausencia de artefactos de compresión, evaluación mensual","Tiempo de almacenamiento disponible antes de sobrescritura, meta mantener mínimo 30 días de historial continuo","Índice de satisfacción de usuarios operadores del CCO, medición trimestral mediante encuesta","Número de eventos de ciberseguridad detectados y mitigados en el sistema de videovigilancia, seguimiento continuo","Costo operativo por kilómetro de vía monitoreada, optimización continua del presupuesto de operación","Reducción porcentual del tiempo de respuesta ante incidentes comparado con el período pre-implementación"]
    },
    "CC-02": {
        "alcance": "Implementación del sistema de Detección Automática de Incidentes (AID) basado en análisis de video y sensores de tráfico, cubriendo los 306 km del corredor Chía-Girardot, con capacidad de detección en tiempo real de eventos críticos como vehículos detenidos, peatones en vía, circulación en contravía, pérdida de carga, incendios y formación de hielo.",
        "descripcion": "Sistema AID que procesa en tiempo real las imágenes de las cámaras CCTV utilizando algoritmos de inteligencia artificial y machine learning para detectar automáticamente incidentes viales. Integra sensores de tráfico inductivos, radares y cámaras térmicas para validación cruzada de eventos. El sistema analiza patrones de flujo vehicular, velocidades promedio y comportamiento anómalo para identificar situaciones de riesgo.",
        "justificacion": "Las zonas de neblina persistente del corredor causan múltiples accidentes por baja visibilidad. La detección automática reduce drásticamente el tiempo entre la ocurrencia de un incidente y su reporte, permitiendo respuesta temprana de cuerpos de socorro y previniendo cadenas de colisiones secundarias.",
        "em": ["Tasa de detección de incidentes mayor al 95% con tasa de falsos positivos inferior al 5%","Tiempo de detección desde la ocurrencia del incidente hasta la alerta inferior a 10 segundos","Capacidad de detección en condiciones de baja visibilidad con cobertura de neblina densa hasta 30 metros","Algoritmos certificados para detección de: vehículos detenidos, peatones, circulación en contravía, pérdida de carga, incendios, humo, formación de hielo","Validación cruzada mediante al menos dos fuentes de detección independientes antes de generar alerta crítica","Capacidad de clasificación automática de incidentes por severidad: baja, media, alta y crítica","Procesamiento de video en el borde (edge computing) para reducir latencia y dependencia de conectividad","Integración nativa con CCTV, DMS/VMS y ECS para respuesta coordinada ante incidentes detectados","Actualización continua de modelos de IA mediante aprendizaje federado con datos anonimizados","Cumplimiento con normas internacionales de sistemas ITS y protocolos de emergencia vial"],
        "rf": ["Interfaz de operador en el CCO que muestre alertas con ubicación GPS exacta, imagen del incidente y clasificación de severidad","Generación automática de tickets de incidente con información pre-llenada para despacho de unidades de emergencia","Capacidad de configuración de zonas de detección personalizadas por tramo con umbrales ajustables","Sistema de aprendizaje continuo que mejore la precisión de detección basado en retroalimentación de operadores","Dashboard de estadísticas de incidentes con análisis de patrones temporales, espaciales y de severidad","Funcionalidad de simulación de incidentes para entrenamiento de operadores sin afectar el sistema productivo","Exportación de datos de incidentes en formatos estándar para análisis externo e integración con sistemas de terceros","Gestión de mantenimiento predictivo del propio sistema AID basado en métricas de rendimiento","Notificaciones automáticas a personal autorizado vía SMS, email y push notifications ante incidentes críticos","Auditoría completa de todas las detecciones con registro de confianza del algoritmo y acciones del operador"],
        "rnf": ["Latencia end-to-end desde detección hasta alerta en pantalla inferior a 3 segundos en condiciones normales de red","Capacidad de procesamiento de video de al menos 500 cámaras simultáneas sin pérdida de frames","Disponibilidad del sistema superior al 99.9% con arquitectura redundante de servidores de análisis","Escalabilidad horizontal para incorporar nuevos algoritmos de detección sin rediseño de arquitectura","Seguridad de datos con anonimización de rostros y placas en el procesamiento de video","Eficiencia energética del sistema de procesamiento, optimización del consumo por inferencia","Facilidad de mantenimiento con actualizaciones de modelos de IA sin detener el servicio","Cumplimiento con regulaciones de protección de datos personales y privacidad en video","Tolerancia a fallos con conmutación automática a servidores de respaldo ante fallas","Compatibilidad con estándares de video ONVIF, RTSP y H.265 para integración con cualquier cámara"],
        "kpi": ["Precisión de detección: porcentaje de incidentes detectados correctamente versus total de incidentes reales","Tasa de falsos positivos: alertas generadas sin incidente real, meta inferior al 5%","Tiempo promedio de detección desde ocurrencia hasta alerta visual en el CCO, meta inferior a 10 segundos","Tasa de validación por operadores: porcentaje de alertas confirmadas como incidentes reales por personal humano","Reducción del tiempo de respuesta de emergencias gracias a detección temprana, meta reducción del 40%","Número de incidentes prevenidos mediante detección de condiciones de riesgo antes de su materialización","Eficiencia computacional: número de cámaras procesadas por servidor de análisis, optimización continua","Índice de satisfacción de operadores del CCO con el sistema AID, encuesta mensual","Ahorro económico estimado por reducción de severidad de incidentes detectados tempranamente","Mejora continua del modelo: incremento porcentual de precisión mes a mes mediante aprendizaje"]
    },
}

print("Generando Fichas ITS - Corredor Chía-Girardot")
print("=" * 50)

wb = openpyxl.Workbook()
wb.remove(wb.active)

header_font = Font(bold=True, size=11)
normal_font = Font(size=10)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for codigo, nombre, desc in SISTEMAS:
    ws = wb.create_sheet(title=codigo)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 120
    
    ws['A1'] = f"FICHA TÉCNICA ITS - CORREDOR CHÍA-GIRARDOT"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    ws['A2'] = ""
    
    ws['A3'] = "CAMPO"
    ws['B3'] = "DESCRIPCIÓN"
    ws['A3'].font = header_font_white
    ws['B3'].font = header_font_white
    ws['A3'].fill = header_fill
    ws['B3'].fill = header_fill
    ws['A3'].border = thin_border
    ws['B3'].border = thin_border
    
    data = CONTENIDO.get(codigo, {})
    
    ws['A4'] = "Código"
    ws['B4'] = f"{codigo}-{nombre}"
    ws['A5'] = "Alcance"
    ws['B5'] = data.get("alcance", f"Sistema de {desc} para el corredor vial Chía-Girardot (306 km)")
    ws['A6'] = "Descripción Funcional / Técnica"
    ws['B6'] = data.get("descripcion", f"Sistema {nombre} diseñado para las condiciones del corredor Chía-Girardot")
    ws['A7'] = "Justificación Estratégica"
    ws['B7'] = data.get("justificacion", f"El sistema {nombre} es crítico para la gestión inteligente del corredor")
    
    em_list = data.get("em", [f"Especificación Mínima {i+1} para el sistema {nombre}" for i in range(10)])
    rf_list = data.get("rf", [f"Requisito Funcional {i+1} para el sistema {nombre}" for i in range(10)])
    rnf_list = data.get("rnf", [f"Requisito No Funcional {i+1} para el sistema {nombre}" for i in range(10)])
    kpi_list = data.get("kpi", [f"Indicador de Desempeño {i+1} para el sistema {nombre}" for i in range(10)])
    
    for i in range(10):
        ws[f'A{8+i}'] = f"EM-{i+1:02d}"
        ws[f'B{8+i}'] = em_list[i] if i < len(em_list) else f"EM-{i+1:02d}"
    for i in range(10):
        ws[f'A{18+i}'] = f"RF-{i+1:02d}"
        ws[f'B{18+i}'] = rf_list[i] if i < len(rf_list) else f"RF-{i+1:02d}"
    for i in range(10):
        ws[f'A{30+i}'] = f"RNF-{i+1:02d}"
        ws[f'B{30+i}'] = rnf_list[i] if i < len(rnf_list) else f"RNF-{i+1:02d}"
    for i in range(10):
        ws[f'A{42+i}'] = f"KPI-{i+1:02d}"
        ws[f'B{42+i}'] = kpi_list[i] if i < len(kpi_list) else f"KPI-{i+1:02d}"
    
    for row in range(4, 54):
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    
    print(f"Generada: {codigo} - {nombre}")

output_path = "/root/.openclaw/workspace/Fichas ITS/Fichas_ITS_Completo.xlsx"
wb.save(output_path)
print(f"\nArchivo guardado: {output_path}")
print(f"Total de hojas: {len(SISTEMAS)}")
